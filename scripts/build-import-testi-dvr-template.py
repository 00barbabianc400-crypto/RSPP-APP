"""Genera templates/import-testi-dvr.xlsx con fogli DATI, _Catalogo, _Liste."""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "templates" / "import-testi-dvr.xlsx"

RISCHI = [
    ("S01", "Incendio ed esplosione"),
    ("S02", "Macchine e attrezzature da lavoro"),
    ("S03", "Impianti ed apparecchiature elettriche"),
    ("S04", "Fulminazione / scariche atmosferiche"),
    ("S05", "Atmosfere esplosive (ATEX)"),
    ("S06", "Ambienti confinati / sospetti inquinamento"),
    ("S07", "Luoghi di lavoro — strutture e ambienti"),
    ("S08", "Rischio taglio e punta"),
    ("S09", "Caduta dall'alto / attivita in quota"),
    ("S10", "Incidenti stradali / circolazione veicoli"),
    ("S11", "Gestione lavoratori disabili"),
    ("F01", "Rumore"),
    ("F02", "Vibrazioni mano-braccio"),
    ("F03", "Vibrazioni corpo intero"),
    ("F04", "Campi elettromagnetici"),
    ("F05", "Radiazioni ottiche artificiali"),
    ("F06", "Radiazioni ionizzanti / Radon"),
    ("F07", "Microclima"),
    ("F08", "Illuminamento (igiene)"),
    ("E01", "Movimentazione manuale dei carichi (MMC)"),
    ("E02", "Sovraccarico biomeccanico arti superiori"),
    ("E03", "Videoterminali (VDT)"),
    ("E04", "Posture incongrue / rischio posturale"),
    ("C01", "Agenti chimici"),
    ("C02", "Agenti cancerogeni e mutageni"),
    ("C03", "Amianto"),
    ("C04", "Gas Radon"),
    ("C05", "Fumo passivo"),
    ("B01", "Agenti biologici generici"),
    ("B02", "Legionellosi"),
    ("B03", "Rischio biologico SARS-CoV-2 / pandemia"),
    ("P01", "Stress lavoro-correlato"),
    ("P02", "Mobbing e burn-out"),
    ("P03", "Rischio rapina e aggressione"),
    ("P04", "Lavoratrici gestanti e madri"),
    ("P05", "Tipologia contrattuale (lavoro atipico)"),
    ("P06", "Invecchiamento lavorativo"),
    ("P07", "Parita di genere"),
    ("P08", "Lavoro agile / smart working"),
    ("P09", "Lavoratori stranieri"),
]

HEADERS = [
    "titolo_testo",
    "id_rischio",
    "livello",
    "tipo_testo",
    "azienda_origine",
    "priorita_intervento",
    "testo_valutazione",
    "misure_in_atto",
    "misure_programmate",
]

EXAMPLE = [
    "Esempio — Incendio basso",
    "S01",
    "Basso",
    "Standard",
    "",
    "Immediato",
    "Testo valutazione di esempio (sostituire o cancellare la riga).",
    "Misure in atto di esempio.",
    "Misure programmate di esempio.",
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATI"

    hdr_fill = PatternFill("solid", fgColor="0078D4")
    hdr_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font

    for c, val in enumerate(EXAMPLE, 1):
        ws.cell(row=2, column=c, value=val)

    widths = [28, 12, 10, 12, 24, 18, 48, 32, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    cat = wb.create_sheet("_Catalogo")
    cat.append(["id_rischio", "nome_rischio", "etichetta"])
    for code, name in RISCHI:
        cat.append([code, name, f"{code} - {name}"])

    lst = wb.create_sheet("_Liste")
    lst.append(["livello", "tipo_testo", "priorita_intervento"])
    livelli = ["Basso", "Medio", "Alto"]
    tipi = ["Standard", "Custom"]
    priorita = ["Immediato", "Breve termine", "Medio termine", "Lungo termine"]
    n = max(len(livelli), len(tipi), len(priorita), len(RISCHI))
    for i in range(n):
        lst.append([
            livelli[i] if i < len(livelli) else "",
            tipi[i] if i < len(tipi) else "",
            priorita[i] if i < len(priorita) else "",
        ])

    n_rischi = len(RISCHI)
    dv_rischio = DataValidation(
        type="list",
        formula1=f"=_Catalogo!$A$2:$A${n_rischi + 1}",
        allow_blank=False,
    )
    dv_livello = DataValidation(type="list", formula1="=_Liste!$A$2:$A$4", allow_blank=False)
    dv_tipo = DataValidation(type="list", formula1="=_Liste!$B$2:$B$3", allow_blank=False)
    dv_priorita = DataValidation(
        type="list",
        formula1="=_Liste!$C$2:$C$5",
        allow_blank=True,
    )
    for dv in (dv_rischio, dv_livello, dv_tipo, dv_priorita):
        ws.add_data_validation(dv)

    for row in range(2, 502):
        dv_rischio.add(f"B{row}")
        dv_livello.add(f"C{row}")
        dv_tipo.add(f"D{row}")
        dv_priorita.add(f"F{row}")

    ist = wb.create_sheet("ISTRUZIONI")
    lines = [
        "Import Testi DVR — RSPP-APP",
        "",
        "Compila il foglio DATI (una riga = un testo).",
        "Colonna id_rischio: codice catalogo (menu a tendina).",
        "Colonna azienda_origine: solo se tipo_testo = Custom (P.IVA 11 cifre o ragione sociale).",
        "Righe vuote (senza titolo e testo) vengono ignorate.",
        "Elimina la riga di esempio prima dell'import se non serve.",
    ]
    for i, line in enumerate(lines, 1):
        ist.cell(row=i, column=1, value=line)

    cat.sheet_state = "hidden"
    lst.sheet_state = "hidden"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Scritto {OUT}")


if __name__ == "__main__":
    main()
